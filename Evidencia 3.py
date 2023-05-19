import sqlite3 
from sqlite3 import Error
import os
import openpyxl
import csv

libros=dict()

try:
    with open("libros.csv",'r', newline="") as archivo:
            lector = csv.reader(archivo)
            next(lector)
            for identificador,titulo,autor, genero,añoPublic,isbn,fechaAdqui in lector:
                libros[int(identificador)]=(titulo,autor,genero,añoPublic,isbn,fechaAdqui)
except:
    print()
    print('No se ha encontrado ningun archivo previo de guardado')
    print('**SE HA GENERADO UN ARCHIVO CSV EN BLANCO**')
    archivo = open("libros.csv","w", newline="")
    grabador = csv.writer(archivo)
    grabador.writerow(("Identificador", "titulo", "autor",'genero','año publicacion',"isbn",'fecha adquisicion'))
    archivo.close()

def RegistrarNuevoEjempar():
    while True:
        global identificador
        print("")
        print("****Registrar Libro*****")
        if libros.keys():
            identificador=max(libros.keys())+1
        else:
            identificador=1
        print()
        titulo=input("Ingresa el titulo: ")
        titulo=titulo.upper()
        
        autor=input("Ingresa el autor: ")
        autor=autor.upper()
        
        genero=input("Ingresa el genero: ")
        genero=genero.upper()
        
        añoPublic=input("Ingresa el año de publicacion: ")
        
        isbn=input("Ingresa el ISBN: ")
        isbn=isbn.upper()

        fechAdqui=input("Ingresa el año de adquisicion: ")
        
        
        ejemplar=(titulo,autor,genero,añoPublic,isbn,fechAdqui)
        
        
        libros[identificador]=ejemplar

        agregar=input("Desea agregar otro libro? [S/N]: ")
        agregar=agregar.upper()
        
        if agregar=="S":
            pass
        elif agregar=="N":
            archivo = open("libros.csv","w", newline="")
            grabador = csv.writer(archivo)
            grabador.writerow(('identificador','titulo','autor','genero','año publicacion','isbn','fecha adquisicion'))
            grabador.writerows([(identificador, datos[0],datos[1],datos[2],datos[3],datos[4],datos[5])for identificador, datos in libros.items()])
            archivo.close()

            break
        else:
            print("opcion no valida")

def MostrarCatalagoCompleto():
    print()
    print("*******Catalago completo********")
    print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
    for libro in libros.values(): 
        
      print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
    opcion=input("\nDesea exportar a csv [C]\nDesas exportar a excel [E]\nNo exportar [N]\n ")
    if opcion=="C":
        archivo = open("Reporte ejemplares existentes.csv","w", newline="")
        grabador = csv.writer(archivo)
        grabador.writerow(('titulo','autor','genero','año publicacion','isbn','fecha adquisicion'))
        grabador.writerows([libro[0],libro[1],libro[2],libro[3],libro[4],libro[5]]for libro in libros.values())
        archivo.close()
    elif opcion=="E":
        libro_wb = openpyxl.Workbook() 
        libro_hoja = libro_wb.active 
        libro_hoja.title = "Libros" 
        
        libro_hoja["A1"] = "Titulo"
        libro_hoja["B1"] = "Autor"
        libro_hoja["C1"] = "Genero"
        libro_hoja["D1"] = "Año Publicacion"
        libro_hoja["E1"] = "ISBN"
        libro_hoja["F1"] = "Fecha Adquisicion"
       
        row = 2
        for libro in libros.values():
            libro_hoja.cell(row=row, column=1, value=libro[0])
            libro_hoja.cell(row=row, column=2, value=libro[1])
            libro_hoja.cell(row=row, column=3, value=libro[2])
            libro_hoja.cell(row=row, column=4, value=libro[3])
            libro_hoja.cell(row=row, column=5, value=libro[4])
            libro_hoja.cell(row=row, column=6, value=libro[5])
            row += 1
        libro_wb.save("Reporte ejemplares existentes.xlsx")

def ReportePorAutor():
    autores=list()
    print()
    print("*******Reporte por autor******")
    autores=list()
    for libro in libros.values():
        autores.append(libro[1])
        for autor in autores:
            if autores.count(autor)>1:
                autores.remove(autor)
    print("---AUTORES DISPONIBLES----")
    for autor in autores:
        print(f"-{autor}")
    
    autor=input("Ingrese el nombre del autor: ")
    autorBuscado=autor.upper()
    try:
        print()
        print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
        for libro in libros.values():
            if libro[1]==autorBuscado:
                print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
                opcion=input("\nDesea exportar a csv [C]\nDesas exportar a excel [E]\nNo exportar [N]\n ")
                if opcion=="C":
                  archivo = open("Reporte autores.csv","w", newline="")
                  grabador = csv.writer(archivo)
                  grabador.writerow(('titulo','autor','genero','año publicacion','isbn','fecha adquisicion'))
                  grabador.writerows([libro[0],libro[1],libro[2],libro[3],libro[4],libro[5]]for libro in libros.values() if libro[1]==autorBuscado)
                  archivo.close()
                if opcion=="E":
                  libro_excel = openpyxl.Workbook()
                  hoja = libro_excel["Sheet"] 
                  hoja.title = ("f{autorBuscado}")
                  hoja["A1"].value = "Titulo"
                  hoja["B1"].value = "Autor"
                  hoja["C1"].value = "Genero"
                  hoja["D1"].value = "Año Publicacion"
                  hoja["E1"].value = "ISBN"
                  hoja["F1"].value = "Fecha Adquisicion"
                  renglon=2
                  hoja.cell(row=renglon, column=1).value = libro[0]  
                  hoja.cell(row=renglon, column=2).value = libro[1]  
                  hoja.cell(row=renglon, column=3).value = libro[2]  
                  hoja.cell(row=renglon, column=4).value = libro[3]  
                  hoja.cell(row=renglon, column=5).value = libro[4] 
                  hoja.cell(row=renglon, column=6).value = libro[5]  
                  libro_excel.save(f"Reporte de {autorBuscado}.xlsx")
        else:
          pass
    except:
        pass

def ReportePorGenero():
    print()
    print("*******Reporte por genero******")

    generos=list()
    for libro in libros.values():
        generos.append(libro[2])
        for genero in generos:
            if generos.count(genero)>1:
                generos.remove(genero)
    print("---GENEROS DISPONIBLES----")
    for genero in generos:
        print(f"-{genero}")

    genero=input("Ingrese el genero: ")
    generoBuscado=genero.upper()
    try:
        print()
        print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
        for libro in libros.values():
            if libro[2]==generoBuscado:
                print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
                opcion=input("\nDesea exportar a csv [C]\nDesas exportar a excel [E]\nNo exportar [N]\n ")
                if opcion=="C":
                    archivo = open("Reporte generos.csv","w", newline="")
                    grabador = csv.writer(archivo)
                    grabador.writerow(('titulo','autor','genero','año publicacion','isbn','fecha adquisicion'))
                    grabador.writerows([libro[0],libro[1],libro[2],libro[3],libro[4],libro[5]]for libro in libros.values() if libro[2]==generoBuscado)
                    archivo.close()
                if opcion=="E":
                  libro_excel = openpyxl.Workbook()
                  hoja = libro_excel["Sheet"] 
                  hoja.title = (f"{generoBuscado}")
                  hoja["A1"].value = "Titulo"
                  hoja["B1"].value = "Autor"
                  hoja["C1"].value = "Genero"
                  hoja["D1"].value = "Año Publicacion"
                  hoja["E1"].value = "ISBN"
                  hoja["F1"].value = "Fecha Adquisicion"
                  renglon=2
                  hoja.cell(row=renglon, column=1).value = libro[0]  
                  hoja.cell(row=renglon, column=2).value = libro[1]  
                  hoja.cell(row=renglon, column=3).value = libro[2]  
                  hoja.cell(row=renglon, column=4).value = libro[3]  
                  hoja.cell(row=renglon, column=5).value = libro[4]  
                  hoja.cell(row=renglon, column=6).value = libro[5] 
                  libro_excel.save(f"Reporte de{generoBuscado}.xlsx")
                else:
                  pass
    except:
        pass

def ReportePorAño():
    print()
    print("*******Reporte por año de publicacion******")
    
    
    años=list()
    for libro in libros.values():
        años.append(libro[3])
        for año in años:
            if años.count(año)>1:
                años.remove(año)
    print("---GENEROS DISPONIBLES----")
    for año in años:
        print(f"-{año}")

    
    año=input("Ingrese el año: ")
    añoBuscado=año.upper()
    try:
        print()
        print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
        for libro in libros.values():
            if libro[3]==añoBuscado:
                print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
                opcion=input("\nDesea exportar a csv [C]\nDesas exportar a excel [E]\nNo exportar [N]\n ")
                if opcion=="C":
                    archivo = open("Reporte año.csv","w", newline="")
                    grabador = csv.writer(archivo)
                    grabador.writerow(('titulo','autor','genero','año publicacion','isbn','fecha adquisicion'))
                    grabador.writerows([libro[0],libro[1],libro[2],libro[3],libro[4],libro[5]]for libro in libros.values() if libro[3]==añoBuscado)
                    archivo.close()
                if opcion=="E":
                  libro_excel = openpyxl.Workbook()
                  hoja = libro_excel["Sheet"] 
                  hoja.title = (f"{añoBuscado}") 
                  hoja["A1"].value = "Titulo"
                  hoja["B1"].value = "Autor"
                  hoja["C1"].value = "Genero"
                  hoja["D1"].value = "Año Publicacion"
                  hoja["E1"].value = "ISBN"
                  hoja["F1"].value = "Fecha Adquisicion"
                  renglon=2
                  hoja.cell(row=renglon, column=1).value = libro[0] 
                  hoja.cell(row=renglon, column=2).value = libro[1]  
                  hoja.cell(row=renglon, column=3).value = libro[2]  
                  hoja.cell(row=renglon, column=4).value = libro[3]  
                  hoja.cell(row=renglon, column=5).value = libro[4] 
                  hoja.cell(row=renglon, column=6).value = libro[5]  
                  libro_excel.save(f"Reporte {añoBuscado}.xlsx")
                else:
                  pass
    except:
        pass

def Reportes():
    while True:
        print()
        print("*****Reportes*****")
        print("*1* Catalago completo")
        print("*2* Reporte por autor")
        print('*3* Reporte por genero')
        print('*4* Reporte por año de publicacion')
        print('*5* Regresar al menu anterior')
        try:
          eleccion=int(input("Selecciona una opcion: "))
          if eleccion==1:
            MostrarCatalagoCompleto()
          if eleccion==2:
            ReportePorAutor()
          if eleccion==3:
            ReportePorGenero()
          if eleccion==4:
            ReportePorAño()
          if eleccion==5:
              break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")
          
def BusquedaPorTitulo():
    print()
    print("*****Busqueda por titulo******")
    titulos=list()
    for libro in libros.values():
        titulos.append(libro[0])
        for titulo in titulos:
            if titulos.count(titulo)>1:
                titulos.remove(titulo)
    print("-----TITULOS DISPONIBLES-----")
    for titulo in titulos:
        print(f"-{titulo}")
    #Consulta
    titulo=input("Ingrese el titulo del libro: ")
    tituloBuscado=titulo.upper()
    try:
        print()
        print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
        for libro in libros.values():
            if libro[0]==tituloBuscado:
                print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
    except:
        pass

def BusquedaPorISBN():
    print()
    print("*****Busqueda por ISBN******")

   
    isbn=input("Ingrese el ISBN del libro: ")
    try:
        print()
        print(f"{'Titulo':15}|{'Autor':20}|{'Genero':10}|{'Año Publicacion':<18}|{'ISBN':13}|{'AÑO Adquisicion'}")
        for libro in libros.values():
            if libro[4]==isbn:
                print(f"{libro[0]:15}|{libro[1]:<20}|{libro[2]:<10}|{libro[3]:<18}|{libro[4]:<13}|{libro[5]} ")
    except:
        pass

def TituloYIsbn():
    while True:
        print()
        print("**********Consulta por titulo y ISBN****")
        print()
        print("*1* busqueda por Titulo")
        print("*2* Busqueda por ISBN")
        print("*3* Volver al menu principal")
        try:
          eleccion=int(input("Elige una opcion: "))
          if eleccion==1:
            BusquedaPorTitulo()
          if eleccion==2:
            BusquedaPorISBN()
          if eleccion==3:
            break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")

def ConsultaYReportes():
    while True:
        print()
        print("*********CONSULTA Y REPORTES****************")
        
        print("*1* Consulta de titulo y ISBN")
        print("*2* Reportes")
        print("*3* Volver al menu principal")
        try:
          eleccion=int(input("Elige una opcion: "))
          if eleccion==1:
            TituloYIsbn()
          if eleccion==2:
            Reportes()
          if eleccion==3:
              break
          else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
        except Exception:
          print("\nDebes ingresar un valor entero. Por favor inténtalo de nuevo.")

def tabla_autor():
  try:
      with sqlite3.connect("Autores.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS autor (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellido TEXT NOT NULL);")
          print("Tabla creada exitosamente")
  except Error as ex:
      print (ex)
  except Exception as ex:
      print(ex)

  nombre = input("Dime el nombre del autor: ")
  apellido = input("Dime el apellido del autor ")
  try:
      with sqlite3.connect("Autores.db") as conn:
          mi_cursor = conn.cursor()
          valores = (nombre, apellido)
          mi_cursor.execute("INSERT INTO autor (nombre, apellido) \
          VALUES(?,?)", valores)
          print(f"La clave asignada fue {mi_cursor.lastrowid}")
  except Error as ex:
      print (ex)
  except Exception as ex:
      print(ex)

def tabla_genero():
  try:
      with sqlite3.connect("Generos.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS genero (clave INTEGER PRIMARY KEY, genero TEXT NOT NULL);")
          print("Tabla creada exitosamente")
  except Error as ex:
      print (ex)
  except Exception as ex:
      print(ex)

  genero= input("Dime el genero: ")

  try:
      with sqlite3.connect("Generos.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("INSERT INTO genero (genero) VALUES(?)", (genero,))
          print(f"La clave asignada fue {mi_cursor.lastrowid}")
  except Error as ex:
        print (ex)
  except Exception as ex:
        print(ex)

def Menu():
    while True:
        print()
        print("***********BIBLIOTECA*************")
        print()
        print("*1* Registrar nuevo ejemplar")
        print("*2* Consultas y reportes")
        print("*3* Registrar autor")
        print("*4* Registrar genero")
        print("*5* Salir")
        print()
        var_elect=input("Ingrese un numero: ")
        if var_elect=="1":
            RegistrarNuevoEjempar()
        if var_elect=="2":
            ConsultaYReportes()
        elif var_elect=="3":
            tabla_autor()
            try:
                with sqlite3.connect("Autores.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM autor ORDER BY nombre")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tnombre\tApellido")
                    print("*" * 30)
                    for clave,nombre,apellido in registros:
                      print(f"{clave:^6}\t{nombre:<10}\t{apellido:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
        elif var_elect=="4":
            tabla_genero()
            try:
                with sqlite3.connect("Generos.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM genero ORDER BY genero")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tGenero")
                    print("*" * 30)
                    for clave,genero, in registros:
                      print(f"{clave:^6}\t{genero:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
            try:
                with sqlite3.connect("Generos.db") as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute("SELECT * FROM genero ORDER BY genero")
                  registros = mi_cursor.fetchall()
                  if registros:
                    print("Claves\tGenero")
                    print("*" * 30)
                    for clave,genero in registros:
                      print(f"{clave:^6}\t{genero:<10}")
                  else:
                        print("No se encontraron registros en la respuesta")
            except Error as ex:
                print (ex)
            except Exception:
                print (ex)
        elif var_elect=="5":
            print("Ha salido del programa")
            break
        else:
            print("\nOpción inválida. Por favor eliga una opción válida.")
Menu()
